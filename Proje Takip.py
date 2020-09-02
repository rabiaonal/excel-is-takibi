import tkinter as tk
import tkinter.filedialog as fl
import pandas as pd 
from datetime import datetime
import shutil 
import openpyxl
import math
from openpyxl.styles.borders import Border, Side



def helloCallBack():
   global file_path
   file_path = fl.askopenfilename(initialdir = "/",title = "Dosya Seç",filetypes=[("Excel files", "*.xlsx")])


window = tk.Tk()
window.title("Proje Takip")
window.geometry("400x300")
greeting = tk.Label(window, justify="left", text="\n\n- Üzerinde çalışılacak excel dosyasını seçtikten sonra\nbu pencereyi kapatınız.\n\n-Oluşturulan kopya orijinal dosyayla aynı yere ismin\nsonuna \"Copy File\" ibaresi eklenerek kaydedilir.\n\n-Daha önceden ortalama gecikme oranı hesaplanmış\ndosya tekrar seçilmeden önce ortalama hücresinin\ntemizlenmesi gereklidir.\n\n")
greeting.pack()


button = tk.Button(
    text="Dosya Seç",
    width=10,
    height=2,
    bg="darkgray",
    fg="black",
    command = helloCallBack
)

button.pack()
window.mainloop()

states = pd.read_excel(file_path)
data = states.values.tolist()

estimate = []
totalWorkDays = []
for i in range(len(data)):
    estimate.append(data[i][9])
 
for st in estimate:
    if type(st) == str:
        WorkDays = 0
        listOfStrings = str(st).split()
        for s in listOfStrings:
            if s[-1] == "w":
                WorkDays += int(s[:-1]) * 7
            elif s[-1] == "d":
                WorkDays += int(s[:-1])
            elif s[-1] == "h":
                WorkDays += 1
        totalWorkDays.append(WorkDays)
    else:
        totalWorkDays.append(0)
        

file_path_copy = file_path[:-5] + "_CopyFile.xlsx"
shutil.copyfile(file_path, file_path_copy)

theFile = openpyxl.load_workbook(file_path_copy)
currentSheet = theFile[theFile.sheetnames[0]]
today = datetime.today()
kalanIs = []
endDate = []
startDate = []
toplam = 0.0

for i in range(len(data)):
    kalanIs.append(math.ceil(  totalWorkDays[i] * (1 - float(data[i][5])) )  )

for i in range(len(data)):
    if len(str(data[i][8])) > 10:
        endDate.append(str(data[i][8]))
    else:
        endDate.append(str(data[i][8]) + " 00:00:00")
    if len(str(data[i][7])) > 10:
        startDate.append(str(data[i][7]))
    else:
        startDate.append(str(data[i][7]) + " 00:00:00")

for i in range(len(kalanIs)):
    if float(data[i][5]) == 1.0:
        currentSheet.cell(i+2 , 7).fill = openpyxl.styles.PatternFill(start_color=("00FF00"), end_color=("00FF00"), fill_type = "solid")
        currentSheet.cell(i+2 , 7).value = "TAMAMLANDI"
        currentSheet.cell(i+2, 13).value = "0.0 %"
    elif int(kalanIs[i]) <= (datetime.strptime(endDate[i] , '%Y-%m-%d %H:%M:%S') - today).days + 2:
        currentSheet.cell(i+2 , 7).fill = openpyxl.styles.PatternFill(start_color=("00FF00"), end_color=("00FF00"), fill_type = "solid")
        currentSheet.cell(i+2 , 7).value = "ZAMANINDA"
        currentSheet.cell(i+2, 13).value = "0.0 %"
    elif datetime.strptime(endDate[i] , '%Y-%m-%d %H:%M:%S').date() < today.date() and float(data[i][5]) < 1.0:
        currentSheet.cell(i+2 , 7).fill = openpyxl.styles.PatternFill(start_color=("FF0000"), end_color=("FF0000"), fill_type = "solid")
        currentSheet.cell(i+2 , 7).value = "GECİKTİ"
        oran = ((today - datetime.strptime(startDate[i] , '%Y-%m-%d %H:%M:%S')).days + 1 + int(kalanIs[i])) / ( (datetime.strptime(endDate[i] , '%Y-%m-%d %H:%M:%S') - datetime.strptime(startDate[i] , '%Y-%m-%d %H:%M:%S')).days + 1 ) - 1
        currentSheet.cell(i+2, 13).value = str(round(oran * 100, 2)) + " %"
        toplam += oran
    else:
       currentSheet.cell(i+2 , 7).fill = openpyxl.styles.PatternFill(start_color=("FFFF00"), end_color=("FFFF00"), fill_type = "solid")
       currentSheet.cell(i+2 , 7).value = "RİSKLİ"
       oran2 = ((today - datetime.strptime(startDate[i] , '%Y-%m-%d %H:%M:%S')).days + 1 + int(kalanIs[i])) / ( (datetime.strptime(endDate[i] , '%Y-%m-%d %H:%M:%S') - datetime.strptime(startDate[i] , '%Y-%m-%d %H:%M:%S')).days + 1 ) - 1
       currentSheet.cell(i+2, 13).value = str(round(oran2 * 100, 2 )) + " %"
       toplam += oran2
        
currentSheet.cell(len(kalanIs) + 2, 13).value = str(round(toplam / len(kalanIs) * 100, 2 )) + " %"
currentSheet.cell(len(kalanIs) + 2, 13).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
theFile.save(file_path_copy)